from openpyxl import load_workbook
from scrape import scrape_ups

data_file = 'tracking_numbers.xlsx'

wb = load_workbook(data_file, data_only=True)

ws = wb['Sheet1']
all_rows = list(ws.rows)

for row in all_rows[2:]:
    status = scrape_ups(int(row[1].value))
    ws.cell(column=3, row=row[1].row, value=status)

wb.save(filename = data_file)